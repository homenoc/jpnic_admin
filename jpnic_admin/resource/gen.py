import datetime
import tempfile
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter


def adjust_cell(ws):
    max_row = ws.max_row + 1
    max_column = ws.max_column + 1
    for i in range(1, max_column):
        max_cell_length = 0
        for j in range(1, max_row):
            value = ws.cell(column=i, row=j).value
            if type(value) is not str:
                continue
            tmp_max_cell_length = len(value) * 1.5
            if max_cell_length < tmp_max_cell_length:
                max_cell_length = tmp_max_cell_length
        ws.column_dimensions[get_column_letter(i)].width = max_cell_length


class GenFile:
    def __init__(self, path=None):
        dt_now = datetime.datetime.now()
        self.time = dt_now.strftime("%Y%m%d_%H%M%S")
        if path is None:
            raise Exception("path is not found...")
        self.path = path

    def write_html(self, path, html_str):
        with open(path, "w", encoding="utf-8") as f1:
            f1.write(html_str)

    def resource(self, infos=None, response=None):
        base_file_path = self.path + "/"
        excel_filename = self.time + ".xlsx"
        filepaths = [{"path": base_file_path + excel_filename, "filename": excel_filename}]
        if infos is None:
            raise Exception("info is not found...")
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "BASE"
        sheet.cell(column=1, row=1, value="最終更新日")
        sheet.cell(column=2, row=1, value="資源管理者番号")
        sheet.cell(column=3, row=1, value="資源管理者略称")
        sheet.cell(column=4, row=1, value="組織名")
        sheet.cell(column=5, row=1, value="組織名(英語)")
        sheet.cell(column=6, row=1, value="郵便番号")
        sheet.cell(column=7, row=1, value="住所")
        sheet.cell(column=8, row=1, value="住所(英語)")
        sheet.cell(column=9, row=1, value="電話番号")
        sheet.cell(column=10, row=1, value="Fax")
        sheet.cell(column=11, row=1, value="資源管理責任者")
        sheet.cell(column=12, row=1, value="連絡担当窓口")
        sheet.cell(column=13, row=1, value="一般問い合わせ窓口")
        sheet.cell(column=14, row=1, value="資源管理者通知アドレス")
        sheet.cell(column=15, row=1, value="アサインメントウィンドウサイズ")
        sheet.cell(column=16, row=1, value="管理開始時刻")
        sheet.cell(column=17, row=1, value="総利用率")
        sheet.cell(column=18, row=1, value="割当済")
        sheet.cell(column=19, row=1, value="総数")
        sheet.cell(column=20, row=1, value="AD Ratio")
        sheet_row_index = 2

        for info in infos:
            rs_list = info["rs_list"]
            usage = "{:.2%}".format(rs_list.assigned_addr_count / rs_list.all_addr_count)
            # baseシートに追記
            sheet.cell(column=1, row=sheet_row_index, value=rs_list.last_checked_at.strftime("%Y/%m/%d %H:%M:%S"))
            sheet.cell(column=2, row=sheet_row_index, value=rs_list.resource_no)
            sheet.cell(column=3, row=sheet_row_index, value=rs_list.resource_admin_short)
            sheet.cell(column=4, row=sheet_row_index, value=rs_list.org)
            sheet.cell(column=5, row=sheet_row_index, value=rs_list.org_en)
            sheet.cell(column=6, row=sheet_row_index, value=rs_list.post_code)
            sheet.cell(column=7, row=sheet_row_index, value=rs_list.address)
            sheet.cell(column=8, row=sheet_row_index, value=rs_list.address_en)
            sheet.cell(column=9, row=sheet_row_index, value=rs_list.tel)
            sheet.cell(column=10, row=sheet_row_index, value=rs_list.fax)
            sheet.cell(column=11, row=sheet_row_index, value=rs_list.admin_handle)
            sheet.cell(column=12, row=sheet_row_index, value=rs_list.email)
            sheet.cell(column=13, row=sheet_row_index, value=rs_list.common_email)
            sheet.cell(column=14, row=sheet_row_index, value=rs_list.notify_email)
            sheet.cell(column=15, row=sheet_row_index, value=rs_list.assignment_size)
            sheet.cell(column=16, row=sheet_row_index, value=rs_list.start_date.strftime("%Y/%m/%d"))
            sheet.cell(column=17, row=sheet_row_index, value=usage)
            sheet.cell(column=18, row=sheet_row_index, value=rs_list.assigned_addr_count)
            sheet.cell(column=19, row=sheet_row_index, value=rs_list.all_addr_count)
            sheet.cell(column=20, row=sheet_row_index, value=rs_list.ad_ratio)
            sheet_row_index += 1

            # 各シート
            export_html_filename = str(info["asn"]) + "_" + info["name"] + ".html"
            export_html_path = base_file_path + export_html_filename
            filepaths.append({"path": export_html_path, "filename": export_html_filename})
            self.write_html(path=export_html_path, html_str=rs_list.html_source)
            ws = wb.create_sheet(title=str(info["asn"]) + "_" + info["name"])
            ws.cell(column=1, row=1, value="総利用率")
            ws.cell(column=2, row=1, value=usage)
            ws.cell(column=1, row=2, value="割当済")
            ws.cell(column=2, row=2, value=rs_list.assigned_addr_count)
            ws.cell(column=1, row=3, value="総数")
            ws.cell(column=2, row=3, value=rs_list.all_addr_count)
            ws.cell(column=1, row=4, value="AD Ratio")
            ws.cell(column=2, row=4, value=rs_list.ad_ratio)
            ws.cell(column=6, row=1, value="取得日")
            ws.cell(column=6, row=2, value=rs_list.last_checked_at.strftime("%Y/%m/%d %H:%M:%S"))

            ws.cell(column=1, row=8, value="CIDR Block")
            ws.cell(column=2, row=8, value="割振日")
            ws.cell(column=3, row=8, value="利用率")
            ws.cell(column=4, row=8, value="割当済(個)")
            ws.cell(column=5, row=8, value="全体(個)")
            row_index = 9
            for rs_addr_list in info["rs_addr_lists"]:
                total_addr = 2 ** (32 - int(str(rs_addr_list.ip_address).split("/")[1]))
                ws.cell(column=1, row=row_index, value=rs_addr_list.ip_address)
                ws.cell(column=2, row=row_index, value=rs_addr_list.assign_date.strftime("%Y/%m/%d"))
                ws.cell(
                    column=3,
                    row=row_index,
                    value="{:.2%}".format(rs_addr_list.assigned_addr_count / total_addr),
                )
                ws.cell(column=4, row=row_index, value=rs_addr_list.assigned_addr_count)
                ws.cell(column=5, row=row_index, value=total_addr)
                row_index += 1
            adjust_cell(ws)

        # adjust base sheet
        adjust_cell(sheet)
        wb.save(base_file_path + excel_filename)
        wb.close()
        zf = zipfile.ZipFile(response, "w")
        for filepath in filepaths:
            zf.write(filepath["path"], arcname=filepath["filename"])

        return self.time + ".zip"
